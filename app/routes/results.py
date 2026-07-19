import io
import os
import re
from datetime import datetime, timedelta

from flask import (
    Blueprint,
    render_template,
    request,
    abort,
    send_file,
)

from flask_login import (
    login_required,
    current_user,
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)

from app.extensions import db
from app.models.attempt import Attempt
from app.models.user import User
from app.models.exam import Exam
from app.logger import app_logger


results_bp = Blueprint(
    "results",
    __name__
)


# ==========================
# SHARED FILTER LOGIC
# ==========================
#
# Used by the results page AND the Excel export -- centralizing it
# here means the page and its export can never drift out of sync
# with each other (e.g. one supporting a filter the other silently
# ignores).

def build_filtered_attempts_query(search, date_from, date_to):

    query = Attempt.query

    if search:

        like = f"%{search}%"

        query = (
            query
            .join(User, User.id == Attempt.student_id)
            .join(Exam, Exam.id == Attempt.exam_id)
            .filter(
                db.or_(
                    User.student_id.ilike(like),
                    User.full_name.ilike(like),
                    User.organization.ilike(like),
                    User.job_title.ilike(like),
                    Exam.code.ilike(like),
                    Exam.title.ilike(like),
                )
            )
        )

    # Dates come in as "YYYY-MM-DD" strings from an <input type="date">.
    # Invalid/empty values are simply ignored rather than erroring out,
    # so a malformed query string just falls back to "no date filter"
    # instead of crashing the page.
    if date_from:

        try:
            start = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(Attempt.created_at >= start)
        except ValueError:
            pass

    if date_to:

        try:
            # Filtering by "< next day" rather than "<= end of day"
            # correctly includes every attempt made ON date_to,
            # regardless of the time component.
            end = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Attempt.created_at < end)
        except ValueError:
            pass

    return query


def safe_filename(text):
    """Strips anything that isn't filename-safe."""
    return re.sub(r"[^A-Za-z0-9_.-]", "_", text)


# ==========================
# ADMIN RESULT HISTORY
# ==========================

@results_bp.route("/results")
@login_required
def results():

    if current_user.role != "admin":
        abort(403)

    search = request.args.get("q", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    query = build_filtered_attempts_query(search, date_from, date_to)

    attempts = query.order_by(
        Attempt.created_at.desc()
    ).all()

    app_logger.info(
        f"RESULTS VIEWED | "
        f"Admin={current_user.username}"
    )

    return render_template(
        "results.html",
        attempts=attempts,
        search=search,
        date_from=date_from,
        date_to=date_to,
    )


# ==========================
# EXPORT RESULTS TO EXCEL
# ==========================
#
# Exports whatever the admin is currently looking at -- same search
# term and date range as the page, passed through as query params by
# the "Export to Excel" link/button in results.html. No search/date
# filter active means "export everything," same as the page itself.

@results_bp.route("/results/export")
@login_required
def export_results():

    if current_user.role != "admin":
        abort(403)

    search = request.args.get("q", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    query = build_filtered_attempts_query(search, date_from, date_to)

    attempts = query.order_by(
        Attempt.created_at.desc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Student ID",
        "Full Name",
        "Organization",
        "Job Title",
        "Exam Code",
        "Exam Title",
        "Score",
        "Percentage",
        "Status",
        "Completed On",
    ]

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0B3D91",
        end_color="0B3D91",
        fill_type="solid",
    )

    for col_num in range(1, len(headers) + 1):

        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for attempt in attempts:

        ws.append([
            attempt.student.student_id,
            attempt.student.full_name,
            attempt.student.organization or "",
            attempt.student.job_title or "",
            attempt.exam.code,
            attempt.exam.title,
            attempt.score,
            attempt.percentage,
            "PASS" if attempt.passed else "FAIL",
            attempt.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Auto-width columns based on content, so the export doesn't need
    # manual column resizing before it's usable.
    for col_num, header in enumerate(headers, 1):

        max_len = len(header)

        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):

            for cell in row:

                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col_num)].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_parts = ["exam_results"]

    if search:
        filename_parts.append(f"search-{search}")

    if date_from or date_to:
        filename_parts.append(f"{date_from or 'start'}_to_{date_to or 'end'}")

    filename = safe_filename("_".join(filename_parts)) + ".xlsx"

    app_logger.info(
        f"RESULTS EXPORTED | "
        f"Admin={current_user.username} | "
        f"Count={len(attempts)} | "
        f"Search={search or '(none)'} | "
        f"DateFrom={date_from or '(none)'} | "
        f"DateTo={date_to or '(none)'}"
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )


# ==========================
# PDF RESULT CERTIFICATE
# ==========================
#
# One attempt at a time -- reached via a "Certificate" link/button
# next to each row in results.html, so it works identically whether
# that row is currently part of the full list, a search, or a
# date-filtered view (it's always the same table, just filtered
# differently).

@results_bp.route("/results/certificate/<int:attempt_id>")
@login_required
def result_certificate(attempt_id):

    if current_user.role != "admin":
        abort(403)

    attempt = Attempt.query.get_or_404(attempt_id)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=40,
        bottomMargin=40,
        leftMargin=50,
        rightMargin=50,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CertTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#0b3d91"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    subtitle_style = ParagraphStyle(
        "CertSubtitle",
        parent=styles["Normal"],
        fontSize=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#555555"),
        spaceAfter=24,
    )

    footer_style = ParagraphStyle(
        "CertFooter",
        parent=styles["Normal"],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#888888"),
    )

    elements = []

    # Logo is best-effort: if the file isn't found at this path (e.g.
    # a different deployment layout), the certificate still generates
    # fine, just without the image up top.
    logo_path = os.path.join(
        "app", "static", "images", "ecaa_logo.png"
    )

    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=64, height=64))
        elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            "Ethiopian Civil Aviation Authority",
            title_style,
        )
    )

    elements.append(
        Paragraph(
            "Computer Based Testing Platform &mdash; "
            "Examination Result Certificate",
            subtitle_style,
        )
    )

    status_text = "PASS" if attempt.passed else "FAIL"

    status_color = (
        colors.HexColor("#166534")
        if attempt.passed
        else colors.HexColor("#991b1b")
    )

    data = [
        ["Examinee Name", attempt.student.full_name],
        ["Examinee ID", attempt.student.student_id],
        ["Organization", attempt.student.organization or "-"],
        ["Job Title", attempt.student.job_title or "-"],
        ["Examination", f"{attempt.exam.code} \u2014 {attempt.exam.title}"],
        ["Score", f"{attempt.score} ({attempt.percentage}%)"],
        ["Pass Mark", f"{attempt.exam.pass_mark}%"],
        ["Result", status_text],
        ["Date Completed", attempt.created_at.strftime("%d %B %Y, %H:%M")],
    ]

    table = Table(data, colWidths=[150, 300])

    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#0b3d91")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#e5e9f0")),
        ("TEXTCOLOR", (1, 7), (1, 7), status_color),
        ("FONTNAME", (1, 7), (1, 7), "Helvetica-Bold"),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 40))

    elements.append(
        Paragraph(
            f"Generated on {datetime.now().strftime('%d %B %Y, %H:%M')} "
            f"by {current_user.full_name}",
            footer_style,
        )
    )

    elements.append(
        Paragraph(
            "This is a system-generated certificate from the "
            "ECAA CBT Platform.",
            footer_style,
        )
    )

    doc.build(elements)
    buffer.seek(0)

    filename = safe_filename(
        f"certificate_{attempt.student.student_id}_{attempt.exam.code}"
    ) + ".pdf"

    app_logger.info(
        f"CERTIFICATE GENERATED | "
        f"Student={attempt.student.student_id} | "
        f"Exam={attempt.exam.code} | "
        f"By={current_user.username}"
    )

    # as_attachment=False: opens in the browser tab so the admin can
    # preview it, then use the browser's own save/print controls --
    # rather than forcing an immediate download like the Excel export.
    return send_file(
        buffer,
        as_attachment=False,
        download_name=filename,
        mimetype="application/pdf",
    )


# ==========================
# STUDENT RESULT HISTORY
# ==========================

@results_bp.route("/my-results")
@login_required
def student_results():

    if current_user.role != "student":
        abort(403)

    attempts = (
        Attempt.query
        .filter_by(student_id=current_user.id)
        .order_by(Attempt.created_at.desc())
        .all()
    )

    app_logger.info(
        f"MY RESULTS VIEWED | "
        f"Student={current_user.student_id}"
    )

    return render_template(
        "student_results.html",
        attempts=attempts
    )