import io

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    abort,
    send_file,
)

from flask_login import (
    login_required,
    current_user,
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.extensions import db
from app.models.exam import Exam
from app.models.question import Question
from app.logger import app_logger


questions_bp = Blueprint(
    "questions",
    __name__,
)

# Exact header row expected in the question import file, in order.
QUESTION_IMPORT_HEADERS = [
    "Question",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Correct Answer",
    "Marks",
]


# ==========================
# QUESTIONS
# ==========================

@questions_bp.route(
    "/exam/<int:id>/questions",
    methods=["GET", "POST"],
)
@login_required
def exam_questions(id):

    if current_user.role != "admin":
        abort(403)

    exam = Exam.query.get_or_404(id)

    if request.method == "POST":

        answer = request.form.get(
            "answer",
            ""
        ).strip().upper()

        if answer not in ["A", "B", "C", "D"]:

            flash(
                "Correct answer must be A, B, C or D."
            )

            return redirect(
                url_for(
                    "questions.exam_questions",
                    id=id,
                )
            )

        q = Question(
            exam_id=id,
            question_text=request.form.get(
                "question",
                ""
            ).strip(),
            option_a=request.form.get(
                "a",
                ""
            ).strip(),
            option_b=request.form.get(
                "b",
                ""
            ).strip(),
            option_c=request.form.get(
                "c",
                ""
            ).strip(),
            option_d=request.form.get(
                "d",
                ""
            ).strip(),
            correct_answer=answer,
            marks=int(
                request.form.get(
                    "marks"
                )
            ),
        )

        db.session.add(q)
        db.session.commit()

        app_logger.info(
            f"QUESTION CREATED | "
            f"Exam={exam.code} | "
            f"QuestionID={q.id} | "
            f"By={current_user.username}"
        )

        flash(
            "Question added successfully."
        )

        return redirect(
            url_for(
                "questions.exam_questions",
                id=id,
            )
        )

    questions = Question.query.filter_by(
        exam_id=id
    ).order_by(
        Question.id.asc()
    ).all()

    return render_template(
        "questions.html",
        exam=exam,
        questions=questions,
    )


# ==========================
# SEARCH QUESTIONS (ALL EXAMS)
# ==========================
#
# NEW: unlike exam_questions() above, which lists questions for ONE
# exam, this searches question text AND answer options across EVERY
# exam at once. Built as its own route/page rather than a filter on
# exam_questions() since there's no single exam context to scope it
# to -- the admin explicitly wants to find a question wherever it
# lives.

@questions_bp.route(
    "/questions/search",
    methods=["GET"],
)
@login_required
def search_questions():

    if current_user.role != "admin":
        abort(403)

    keyword = request.args.get(
        "q",
        ""
    ).strip()

    results = []

    if keyword:

        like = f"%{keyword}%"

        # Explicit join + tuple pairing rather than relying on a
        # Question.exam relationship (not confirmed to exist on the
        # model) -- same defensive pattern used for the assignment
        # management page.
        rows = (
            db.session.query(Question, Exam)
            .join(Exam, Exam.id == Question.exam_id)
            .filter(
                db.or_(
                    Question.question_text.ilike(like),
                    Question.option_a.ilike(like),
                    Question.option_b.ilike(like),
                    Question.option_c.ilike(like),
                    Question.option_d.ilike(like),
                )
            )
            .order_by(
                Exam.code.asc(),
                Question.id.asc(),
            )
            .all()
        )

        results = [
            {"question": q, "exam": e}
            for q, e in rows
        ]

    return render_template(
        "search_questions.html",
        keyword=keyword,
        results=results,
    )


# ==========================
# EDIT QUESTION
# ==========================

@questions_bp.route(
    "/question/edit/<int:id>",
    methods=["GET", "POST"],
)
@login_required
def edit_question(id):

    if current_user.role != "admin":
        abort(403)

    question = Question.query.get_or_404(id)

    if request.method == "POST":

        answer = request.form.get(
            "answer",
            ""
        ).strip().upper()

        if answer not in ["A", "B", "C", "D"]:

            flash(
                "Correct answer must be A, B, C or D."
            )

            return redirect(
                url_for(
                    "questions.edit_question",
                    id=id,
                )
            )

        question.question_text = request.form.get(
            "question",
            ""
        ).strip()

        question.option_a = request.form.get(
            "a",
            ""
        ).strip()

        question.option_b = request.form.get(
            "b",
            ""
        ).strip()

        question.option_c = request.form.get(
            "c",
            ""
        ).strip()

        question.option_d = request.form.get(
            "d",
            ""
        ).strip()

        question.correct_answer = answer

        question.marks = int(
            request.form.get(
                "marks"
            )
        )

        db.session.commit()

        app_logger.info(
            f"QUESTION UPDATED | "
            f"QuestionID={question.id} | "
            f"By={current_user.username}"
        )

        flash(
            "Question updated successfully."
        )

        return redirect(
            url_for(
                "questions.exam_questions",
                id=question.exam_id,
            )
        )

    return render_template(
        "edit_question.html",
        question=question,
    )


# ==========================
# DELETE QUESTION
# ==========================

@questions_bp.route(
    "/question/delete/<int:id>"
)
@login_required
def delete_question(id):

    if current_user.role != "admin":
        abort(403)

    question = Question.query.get_or_404(id)

    exam_id = question.exam_id

    app_logger.info(
        f"QUESTION DELETED | "
        f"QuestionID={question.id} | "
        f"By={current_user.username}"
    )

    db.session.delete(question)
    db.session.commit()

    flash(
        "Question deleted successfully."
    )

    return redirect(
        url_for(
            "questions.exam_questions",
            id=exam_id,
        )
    )


# ==========================
# DOWNLOAD IMPORT TEMPLATE
# ==========================

@questions_bp.route(
    "/exam/<int:id>/questions/import/template",
    methods=["GET"],
)
@login_required
def download_question_import_template(id):

    if current_user.role != "admin":
        abort(403)

    # Confirms the exam exists (404s otherwise) even though the
    # template itself doesn't depend on which exam it's for -- keeps
    # the URL structure consistent with the rest of this file, and
    # means the download link on the per-exam page never points at a
    # nonexistent exam.
    Exam.query.get_or_404(id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    ws.append(QUESTION_IMPORT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0B3D91",
        end_color="0B3D91",
        fill_type="solid",
    )

    for col_num in range(1, len(QUESTION_IMPORT_HEADERS) + 1):

        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.append([
        "What is the capital of Ethiopia?",
        "Addis Ababa",
        "Nairobi",
        "Cairo",
        "Lagos",
        "A",
        1,
    ])

    ws.column_dimensions["A"].width = 45
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="question_import_template.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )


# ==========================
# IMPORT QUESTIONS FROM EXCEL
# ==========================
#
# Expected columns, in order: Question, Option A, Option B, Option C,
# Option D, Correct Answer, Marks.
#
# Every field is required for a row to import. Correct Answer must be
# A/B/C/D (case-insensitive). Marks must be a positive whole number.
# Invalid rows are skipped with a reason; the rest of the file still
# imports.

@questions_bp.route(
    "/exam/<int:id>/questions/import",
    methods=["GET", "POST"],
)
@login_required
def import_questions(id):

    if current_user.role != "admin":
        abort(403)

    exam = Exam.query.get_or_404(id)

    if request.method == "GET":

        return render_template(
            "import_questions.html",
            exam=exam,
            report=None,
        )

    uploaded_file = request.files.get("import_file")

    if not uploaded_file or uploaded_file.filename == "":

        flash(
            "Please choose an Excel file to import."
        )

        return redirect(
            url_for("questions.import_questions", id=id)
        )

    try:

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

    except Exception:

        flash(
            "Could not read that file. Make sure it's a valid "
            ".xlsx file, ideally based on the downloadable template."
        )

        return redirect(
            url_for("questions.import_questions", id=id)
        )

    report = {
        "created": [],
        "errors": [],
    }

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row_number, row in enumerate(rows, start=2):

        if row is None or all(cell in (None, "") for cell in row):
            continue

        question_text = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        option_a = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        option_b = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        option_c = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        option_d = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        raw_answer = str(row[5]).strip().upper() if len(row) > 5 and row[5] is not None else ""
        raw_marks = row[6] if len(row) > 6 else None

        if not all([question_text, option_a, option_b, option_c, option_d]):

            report["errors"].append({
                "row": row_number,
                "reason": "Question text and all four options are required.",
            })

            continue

        if raw_answer not in ["A", "B", "C", "D"]:

            report["errors"].append({
                "row": row_number,
                "reason": f"Correct Answer must be A, B, C or D (got '{raw_answer}').",
            })

            continue

        try:

            marks = int(raw_marks)

            if marks < 1:
                raise ValueError

        except (TypeError, ValueError):

            report["errors"].append({
                "row": row_number,
                "reason": f"Marks must be a positive whole number (got '{raw_marks}').",
            })

            continue

        question = Question(
            exam_id=id,
            question_text=question_text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            correct_answer=raw_answer,
            marks=marks,
        )

        db.session.add(question)

        report["created"].append({
            "row": row_number,
            "question_text": question_text,
        })

    db.session.commit()

    app_logger.info(
        f"QUESTION IMPORT | "
        f"Exam={exam.code} | "
        f"Created={len(report['created'])} | "
        f"Errors={len(report['errors'])} | "
        f"By={current_user.username}"
    )

    return render_template(
        "import_questions.html",
        exam=exam,
        report=report,
    )