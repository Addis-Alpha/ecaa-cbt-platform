import io
import secrets
import string

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    abort,
    send_file,
)

from flask_login import (
    login_required,
    current_user,
)

from werkzeug.security import (
    generate_password_hash,
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.extensions import db
from app.models.user import User
from app.logger import app_logger, security_logger


students_bp = Blueprint(
    "students",
    __name__,
)


# Exact header row expected in the import file, in this order. The
# template download route produces a file with exactly these headers
# so there's no guesswork for the admin filling it in.
IMPORT_HEADERS = [
    "Student ID",
    "Full Name",
    "Organization",
    "Job Title",
    "Password",
]


# ==========================
# MANAGE STUDENTS
# ==========================

@students_bp.route(
    "/students",
    methods=["GET", "POST"],
)
@login_required
def students():

    if current_user.role != "admin":
        abort(403)

    if request.method == "POST":

        student_id = request.form.get(
            "student_id",
            ""
        ).strip()

        full_name = request.form.get(
            "full_name",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        # NEW: required fields, same as the others -- the model
        # itself doesn't enforce this at the DB level (nullable),
        # so it's checked here, consistent with how the rest of this
        # route already works.
        organization = request.form.get(
            "organization",
            ""
        ).strip()

        job_title = request.form.get(
            "job_title",
            ""
        ).strip()

        if not organization or not job_title:

            flash(
                "Organization and Job Title are required."
            )

            return redirect(
                url_for(
                    "students.students"
                )
            )

        existing = User.query.filter_by(
            student_id=student_id
        ).first()

        if existing:

            flash(
                "Student ID already exists."
            )

        else:

            student = User(
                username=None,
                student_id=student_id,
                full_name=full_name,
                password=generate_password_hash(password),
                role="student",
                organization=organization,
                job_title=job_title,
            )

            db.session.add(student)
            db.session.commit()

            app_logger.info(
                f"STUDENT CREATED | "
                f"StudentID={student.student_id} | "
                f"By={current_user.username}"
            )

            flash(
                "Student added successfully."
            )

            return redirect(
                url_for(
                    "students.students"
                )
            )

    # Search/filter examinees by ID or name.
    # GET /students?q=keyword -- case-insensitive, matches either
    # field. Empty/omitted "q" behaves exactly as before (full list).
    search = request.args.get(
        "q",
        ""
    ).strip()

    query = User.query.filter_by(
        role="student"
    )

    if search:

        like = f"%{search}%"

        query = query.filter(
            db.or_(
                User.student_id.ilike(like),
                User.full_name.ilike(like),
                User.organization.ilike(like),
                User.job_title.ilike(like),
            )
        )

    students = query.order_by(
        User.full_name
    ).all()

    return render_template(
        "students.html",
        students=students,
        search=search,
    )


# ==========================
# EDIT STUDENT
# ==========================

@students_bp.route(
    "/students/edit/<int:id>",
    methods=["GET", "POST"],
)
@login_required
def edit_student(id):

    if current_user.role != "admin":
        abort(403)

    student = User.query.get_or_404(id)

    if request.method == "POST":

        new_student_id = request.form.get(
            "student_id",
            ""
        ).strip()

        # NEW: required, same as create.
        organization = request.form.get(
            "organization",
            ""
        ).strip()

        job_title = request.form.get(
            "job_title",
            ""
        ).strip()

        if not organization or not job_title:

            flash(
                "Organization and Job Title are required."
            )

            return redirect(
                url_for(
                    "students.edit_student",
                    id=id,
                )
            )

        duplicate = User.query.filter(
            User.student_id == new_student_id,
            User.id != student.id,
        ).first()

        if duplicate:

            flash(
                "Student ID already exists."
            )

            return redirect(
                url_for(
                    "students.edit_student",
                    id=id,
                )
            )

        student.student_id = new_student_id
        student.full_name = request.form.get(
            "full_name",
            ""
        ).strip()

        student.organization = organization
        student.job_title = job_title

        password = request.form.get(
            "password",
            ""
        )

        if password:

            student.password = generate_password_hash(
                password
            )

        db.session.commit()

        app_logger.info(
            f"STUDENT UPDATED | "
            f"StudentID={student.student_id} | "
            f"By={current_user.username}"
        )

        flash(
            "Student updated successfully."
        )

        return redirect(
            url_for(
                "students.students"
            )
        )

    return render_template(
        "edit_student.html",
        student=student,
    )


# ==========================
# DELETE STUDENT
# ==========================

@students_bp.route(
    "/students/delete/<int:id>"
)
@login_required
def delete_student(id):

    if current_user.role != "admin":
        abort(403)

    student = User.query.get_or_404(id)

    if student.role != "student":
        abort(403)

    app_logger.info(
        f"STUDENT DELETED | "
        f"StudentID={student.student_id} | "
        f"By={current_user.username}"
    )

    db.session.delete(student)
    db.session.commit()

    flash(
        "Student deleted successfully."
    )

    return redirect(
        url_for(
            "students.students"
        )
    )


# ==========================
# DOWNLOAD IMPORT TEMPLATE
# ==========================

@students_bp.route(
    "/students/import/template",
    methods=["GET"],
)
@login_required
def download_import_template():

    if current_user.role != "admin":
        abort(403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(IMPORT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0B3D91",
        end_color="0B3D91",
        fill_type="solid",
    )

    for col_num in range(1, len(IMPORT_HEADERS) + 1):

        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # One example row so the format is obvious at a glance.
    ws.append([
        "S12345",
        "Jane Doe",
        "Example Organization",
        "Safety Officer",
        "",
    ])

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="student_import_template.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )


# ==========================
# IMPORT STUDENTS FROM EXCEL
# ==========================
#
# Expected columns, in order: Student ID, Full Name, Organization,
# Job Title, Password (optional -- see below).
#
# - Duplicate Student IDs are SKIPPED; the existing record is left
#   untouched.
# - Rows missing a required field (ID, name, organization, job
#   title) are skipped with a reason, the rest of the file still
#   imports.
# - If Password is left blank for a row, a random temporary password
#   is generated and that account is flagged must_change_password
#   (reusing the same mechanism as the default admin account) -- the
#   generated password is shown ONCE in the results report so the
#   admin can hand it to the student, and is never stored anywhere
#   in plaintext.

def generate_temp_password():

    alphabet = string.ascii_letters + string.digits

    return "".join(secrets.choice(alphabet) for _ in range(10))


@students_bp.route(
    "/students/import",
    methods=["GET", "POST"],
)
@login_required
def import_students():

    if current_user.role != "admin":
        abort(403)

    if request.method == "GET":

        return render_template(
            "import_students.html",
            report=None,
        )

    uploaded_file = request.files.get("import_file")

    if not uploaded_file or uploaded_file.filename == "":

        flash(
            "Please choose an Excel file to import."
        )

        return redirect(
            url_for("students.import_students")
        )

    try:

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

    except Exception:

        flash(
            "Could not read that file. Make sure it's a valid "
            ".xlsx file, ideally based on the downloadable template."
        )

        return redirect(
            url_for("students.import_students")
        )

    report = {
        "created": [],
        "skipped": [],
        "errors": [],
    }

    # Skip the header row (row 1), read from row 2 onward.
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row_number, row in enumerate(rows, start=2):

        if row is None or all(cell in (None, "") for cell in row):
            # Blank row -- silently skip, not worth reporting as an
            # error.
            continue

        student_id = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        full_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        organization = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        job_title = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        provided_password = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

        if not student_id or not full_name or not organization or not job_title:

            report["errors"].append({
                "row": row_number,
                "reason": (
                    "Missing required field (Student ID, Full Name, "
                    "Organization, and Job Title are all required)."
                ),
            })

            continue

        existing = User.query.filter_by(
            student_id=student_id
        ).first()

        if existing:

            report["skipped"].append({
                "row": row_number,
                "student_id": student_id,
                "reason": "Student ID already exists.",
            })

            continue

        temp_password = None
        must_change = False

        if provided_password:
            password_to_hash = provided_password
        else:
            temp_password = generate_temp_password()
            password_to_hash = temp_password
            must_change = True

        student = User(
            username=None,
            student_id=student_id,
            full_name=full_name,
            organization=organization,
            job_title=job_title,
            password=generate_password_hash(password_to_hash),
            role="student",
            must_change_password=must_change,
        )

        db.session.add(student)

        report["created"].append({
            "row": row_number,
            "student_id": student_id,
            "full_name": full_name,
            "temp_password": temp_password,
        })

    db.session.commit()

    app_logger.info(
        f"STUDENT IMPORT | "
        f"Created={len(report['created'])} | "
        f"Skipped={len(report['skipped'])} | "
        f"Errors={len(report['errors'])} | "
        f"By={current_user.username}"
    )

    return render_template(
        "import_students.html",
        report=report,
    )


from flask import jsonify


# ==========================
# LIVE SESSION STATUS (AJAX)
# ==========================
#
# NEW: returns current active/inactive status for every student, as
# JSON. Polled periodically by students.html via JavaScript so the
# admin sees live updates (a student logging out, or being force-
# logged-out) without needing to manually refresh the page.

@students_bp.route(
    "/students/status",
    methods=["GET"],
)
@login_required
def students_status():

    if current_user.role != "admin":
        abort(403)

    students = User.query.filter_by(
        role="student"
    ).all()

    data = {
        str(s.id): bool(s.active_session_token)
        for s in students
    }

    return jsonify(data)